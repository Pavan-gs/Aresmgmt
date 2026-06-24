"""
FUND OPERATIONS AUTOMATION WORKSHOP
Sample Data Generator
=====================
Run this script ONCE before the workshop.
Creates everything participants need to run all 15 labs.

Usage:
    python create_workshop_data.py

Output structure:
    E:/aresmgmt/material
    ├── data/
    │   ├── input/
    │   │   ├── format_a_citco/      ← Funds 1-5  (Citco-style format)
    │   │   ├── format_b_ssc/        ← Funds 6-10 (SS&C-style format)
    │   │   ├── format_c_generic/    ← Funds 11-15 (varied formats)
    │   │   ├── pdf_reports/         ← Fund statement text files (simulate PDFs)
    │   │   └── broken/              ← 2 broken files for error-handling labs
    │   ├── reference/               ← fund_master.xlsx, currency_rates.xlsx
    │   └── sql/                     ← funddb.sqlite (portable SQL Server replacement)
    ├── config/                      ← settings.json, exception_rules.json
    ├── logs/
    └── reports/
        ├── daily/
        ├── monthly/
        └── exceptions/
"""

import os, json, sqlite3, random
from pathlib import Path
from datetime import date, datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ─── ROOT (works on Windows + Mac/Linux) ─────────────────────────────────────
ROOT = Path("E:/aresmgmt/material") if os.name == "nt" else Path.home() / "FundAutomation"

# ─── MASTER FUND DEFINITIONS ──────────────────────────────────────────────────
TODAY   = date.today()
VALDATE = TODAY - timedelta(days=1)  # prior business day

FUNDS = [
    # FORMAT A — CITCO STYLE (funds 1-5)
    # Layout: labelled cells, NAV in B5, date in B3, name in A1
    {
        "code": "ALPHA001", "name": "Alpha Capital Equity Fund",
        "admin": "Citco Fund Services", "admin_fmt": "citco",
        "currency": "USD", "nav": 125_432_891.50, "units": 1_000_000.00,
        "mgmt_fee_rate": 0.015, "admin_fee_rate": 0.0025,
        "strategy": "Long/Short Equity", "manager": "Alpha Capital Management",
        "status": "ok", "nav_prior": 124_800_000.00,
    },
    {
        "code": "BETA002", "name": "Beta Growth Fixed Income Fund",
        "admin": "Citco Fund Services", "admin_fmt": "citco",
        "currency": "USD", "nav": 88_210_445.00, "units": 500_000.00,
        "mgmt_fee_rate": 0.010, "admin_fee_rate": 0.002,
        "strategy": "Fixed Income", "manager": "Beta Asset Management",
        "status": "future_date",  # ← intentional: date = tomorrow
        "nav_prior": 88_190_000.00,
    },
    {
        "code": "GAMMA003", "name": "Gamma Absolute Return Fund",
        "admin": "Citco Fund Services", "admin_fmt": "citco",
        "currency": "EUR", "nav": 210_750_000.00, "units": 2_500_000.00,
        "mgmt_fee_rate": 0.020, "admin_fee_rate": 0.003,
        "strategy": "Absolute Return", "manager": "Gamma Partners",
        "status": "ok", "nav_prior": 209_000_000.00,
    },
    {
        "code": "DELTA004", "name": "Delta Multi-Strategy Fund",
        "admin": "Citco Fund Services", "admin_fmt": "citco",
        "currency": "USD", "nav": 450_000_000.00, "units": 3_000_000.00,
        "mgmt_fee_rate": 0.015, "admin_fee_rate": 0.0025,
        "strategy": "Multi-Strategy", "manager": "Delta Capital",
        "status": "large_move",  # ← intentional: NAV moved >8% vs prior
        "nav_prior": 416_000_000.00,
    },
    {
        "code": "EPSILON005", "name": "Epsilon Global Macro Fund",
        "admin": "Citco Fund Services", "admin_fmt": "citco",
        "currency": "GBP", "nav": 175_600_000.00, "units": 1_200_000.00,
        "mgmt_fee_rate": 0.020, "admin_fee_rate": 0.003,
        "strategy": "Global Macro", "manager": "Epsilon Global",
        "status": "ok", "nav_prior": 174_900_000.00,
    },

    # FORMAT B — SS&C STYLE (funds 6-10)
    # Layout: tabular, headers in row 1, data in row 2, different sheet name
    {
        "code": "ZETA006", "name": "Zeta Credit Opportunities Fund",
        "admin": "SS&C GlobeOp", "admin_fmt": "ssc",
        "currency": "USD", "nav": 92_100_000.00, "units": 750_000.00,
        "mgmt_fee_rate": 0.015, "admin_fee_rate": 0.002,
        "strategy": "Credit", "manager": "Zeta Credit Advisors",
        "status": "ok", "nav_prior": 91_800_000.00,
    },
    {
        "code": "ETA007", "name": "Eta Long Short Equity Fund",
        "admin": "SS&C GlobeOp", "admin_fmt": "ssc",
        "currency": "USD", "nav": 310_500_000.00, "units": 2_000_000.00,
        "mgmt_fee_rate": 0.020, "admin_fee_rate": 0.003,
        "strategy": "Long/Short Equity", "manager": "Eta Capital",
        "status": "ok", "nav_prior": 308_000_000.00,
    },
    {
        "code": "THETA008", "name": "Theta Emerging Markets Fund",
        "admin": "SS&C GlobeOp", "admin_fmt": "ssc",
        "currency": "USD", "nav": 155_900_000.00, "units": 1_100_000.00,
        "mgmt_fee_rate": 0.015, "admin_fee_rate": 0.0025,
        "strategy": "Emerging Markets", "manager": "Theta EM Partners",
        "status": "ok", "nav_prior": 155_000_000.00,
    },
    {
        "code": "IOTA009", "name": "Iota Real Asset Fund",
        "admin": "SS&C GlobeOp", "admin_fmt": "ssc",
        "currency": "EUR", "nav": 280_000_000.00, "units": 1_800_000.00,
        "mgmt_fee_rate": 0.010, "admin_fee_rate": 0.002,
        "strategy": "Real Assets", "manager": "Iota Real Assets Ltd",
        "status": "ok", "nav_prior": 279_000_000.00,
    },
    {
        "code": "KAPPA010", "name": "Kappa Quantitative Fund",
        "admin": "SS&C GlobeOp", "admin_fmt": "ssc",
        "currency": "USD", "nav": 420_300_000.00, "units": 3_500_000.00,
        "mgmt_fee_rate": 0.015, "admin_fee_rate": 0.002,
        "strategy": "Quantitative", "manager": "Kappa Quant LLC",
        "status": "ok", "nav_prior": 419_000_000.00,
    },

    # FORMAT C — GENERIC / VARIED (funds 11-15)
    # Sheet names, layouts, and date formats all vary
    {
        "code": "LAMBDA011", "name": "Lambda Infrastructure Fund",
        "admin": "Generic Admin Co", "admin_fmt": "generic",
        "currency": "GBP", "nav": 500_000_000.00, "units": 4_000_000.00,
        "mgmt_fee_rate": 0.010, "admin_fee_rate": 0.0015,
        "strategy": "Infrastructure", "manager": "Lambda Infra Managers",
        "status": "ok", "nav_prior": 499_000_000.00,
        "sheet_name": "Monthly Report",  # varies from NAV Summary
    },
    {
        "code": "MU012", "name": "Mu Healthcare Innovation Fund",
        "admin": "Generic Admin Co", "admin_fmt": "generic",
        "currency": "USD", "nav": 68_450_000.00, "units": 400_000.00,
        "mgmt_fee_rate": 0.020, "admin_fee_rate": 0.003,
        "strategy": "Healthcare", "manager": "Mu Life Sciences Capital",
        "status": "ok", "nav_prior": 67_900_000.00,
        "sheet_name": "NAV Report",
    },
    {
        "code": "NU013", "name": "Nu Technology Growth Fund",
        "admin": "Generic Admin Co", "admin_fmt": "generic",
        "currency": "USD", "nav": 195_700_000.00, "units": 1_500_000.00,
        "mgmt_fee_rate": 0.020, "admin_fee_rate": 0.003,
        "strategy": "Technology", "manager": "Nu Tech Investors",
        "status": "broken_sheet",  # ← intentional: sheet name typo
        "sheet_name": "NAV_Summary",  # underscore not space!
    },
    {
        "code": "XI014", "name": "Xi ESG Sustainability Fund",
        "admin": "Generic Admin Co", "admin_fmt": "generic",
        "currency": "EUR", "nav": 340_200_000.00, "units": 2_200_000.00,
        "mgmt_fee_rate": 0.015, "admin_fee_rate": 0.002,
        "strategy": "ESG", "manager": "Xi Sustainable Investments",
        "status": "ok", "nav_prior": 339_000_000.00,
        "sheet_name": "Fund Data",
    },
    {
        "code": "OMICRON015", "name": "Omicron Distressed Debt Fund",
        "admin": "Generic Admin Co", "admin_fmt": "generic",
        "currency": "USD", "nav": None,  # ← intentional: NAV value missing
        "units": 900_000.00,
        "mgmt_fee_rate": 0.020, "admin_fee_rate": 0.003,
        "strategy": "Distressed Debt", "manager": "Omicron Credit Management",
        "status": "missing_nav",
        "sheet_name": "NAV Summary",
    },
]

# ─── STYLES ──────────────────────────────────────────────────────────────────
def navy_fill():   return PatternFill("solid", fgColor="1A2E4A")
def teal_fill():   return PatternFill("solid", fgColor="0D7C66")
def light_fill():  return PatternFill("solid", fgColor="F0F4F8")
def header_font(): return Font(bold=True, color="FFFFFF", size=11)
def label_font():  return Font(bold=True, color="1A2E4A", size=10)
def data_font():   return Font(color="1A202C", size=10)
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ══════════════════════════════════════════════════════════════════════════════
# FORMAT A — CITCO STYLE
# ══════════════════════════════════════════════════════════════════════════════
def create_citco_file(fund: dict, output_path: Path):
    wb = openpyxl.Workbook()

    # ── NAV Summary sheet ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "NAV Summary"

    # Header band
    for col in range(1, 5):
        ws.cell(1, col).fill = navy_fill()
    ws.merge_cells("A1:D1")
    ws["A1"] = f"CITCO FUND SERVICES  |  NAV REPORT"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Fund name
    ws["A2"] = fund["name"]
    ws["A2"].font = Font(bold=True, size=14, color="1A2E4A")
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 22

    # Spacer
    ws.row_dimensions[3].height = 6

    # Key fields (label in col A, value in col B)
    nav_date = TODAY + timedelta(days=1) if fund["status"] == "future_date" else VALDATE
    fields = [
        ("A4", "Valuation Date:",  "B4", nav_date),
        ("A5", "Net Asset Value:", "B5", fund["nav"]),
        ("A6", "Currency:",        "B6", fund["currency"]),
        ("A7", "Units Outstanding:","B7",fund["units"]),
        ("A8", "NAV per Unit:",    "B8",
            round(fund["nav"] / fund["units"], 4) if fund["nav"] and fund["units"] else None),
        ("A9", "", "B9", ""),
        ("A10","Management Fee Rate:","B10", fund["mgmt_fee_rate"]),
        ("A11","Administration Fee Rate:","B11", fund["admin_fee_rate"]),
        ("A12","Administrator:","B12", fund["admin"]),
        ("A13","Strategy:","B13", fund["strategy"]),
    ]
    for lbl_cell, lbl_val, dat_cell, dat_val in fields:
        if lbl_val:
            ws[lbl_cell] = lbl_val
            ws[lbl_cell].font = label_font()
            ws[lbl_cell].fill = light_fill()
        ws[dat_cell] = dat_val
        ws[dat_cell].font = data_font()
        if isinstance(dat_val, float) and dat_val and dat_val > 1000:
            ws[dat_cell].number_format = '#,##0.00'
        if isinstance(dat_val, date):
            ws[dat_cell].number_format = 'DD-MMM-YYYY'

    # ── Positions sheet ────────────────────────────────────────────────────
    wsp = wb.create_sheet("Positions")
    pos_headers = ["Security ID","Security Name","Asset Class","Quantity","Price","Market Value","Currency","% NAV"]
    for ci, h in enumerate(pos_headers, 1):
        c = wsp.cell(1, ci, h)
        c.font = header_font()
        c.fill = teal_fill()
        c.alignment = Alignment(horizontal="center")
    
    positions = _generate_positions(fund, 8)
    for ri, pos in enumerate(positions, 2):
        for ci, val in enumerate(pos, 1):
            wsp.cell(ri, ci, val)
            if ci in (4, 5, 6):
                wsp.cell(ri, ci).number_format = '#,##0.00'

    set_col_width(ws, "A", 28)
    set_col_width(ws, "B", 22)
    set_col_width(ws, "C", 14)
    wb.save(output_path)
    print(f"  ✓ Created {output_path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# FORMAT B — SS&C STYLE
# ══════════════════════════════════════════════════════════════════════════════
def create_ssc_file(fund: dict, output_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fund_Performance"       # ← different from Citco

    # Row 1 — report title
    ws.merge_cells("A1:J1")
    ws["A1"] = "SS&C GLOBEOP  |  FUND PERFORMANCE REPORT"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=11)
    ws["A1"].fill = navy_fill()
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    # Row 2 — generated timestamp
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  |  Confidential"
    ws["A2"].font = Font(italic=True, color="718096", size=9)

    # Row 3 — column headers (tabular format — different from Citco)
    headers = [
        "Report Date","Fund Code","Fund Name","Currency",
        "Total NAV","Units Outstanding","NAV Per Unit",
        "Management Fee Rate","Admin Fee Rate","Administrator"
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(3, ci, h)
        c.font = header_font()
        c.fill = teal_fill()
        c.alignment = Alignment(horizontal="center")

    # Row 4 — data (all on one row, tabular)
    nav_date = VALDATE
    nav_per_unit = round(fund["nav"] / fund["units"], 4) if fund["nav"] and fund["units"] else None
    data_row = [
        nav_date, fund["code"], fund["name"], fund["currency"],
        fund["nav"], fund["units"], nav_per_unit,
        fund["mgmt_fee_rate"], fund["admin_fee_rate"], fund["admin"]
    ]
    for ci, val in enumerate(data_row, 1):
        c = ws.cell(4, ci, val)
        if isinstance(val, float) and val > 1000:
            c.number_format = '#,##0.00'
        if isinstance(val, date):
            c.number_format = 'DD/MM/YYYY'   # ← different date format from Citco!
        c.border = thin_border()

    # Historical data tab
    wsh = wb.create_sheet("NAV History")
    hist_headers = ["Date","Total NAV","NAV Per Unit","Units Outstanding","Change %"]
    for ci, h in enumerate(hist_headers, 1):
        c = wsh.cell(1, ci, h)
        c.font = header_font()
        c.fill = navy_fill()

    base_nav = fund["nav"] * 0.85
    for i in range(12):
        hist_date = VALDATE - timedelta(days=30 * (11 - i))
        hist_nav = base_nav * (1 + 0.01 * i + random.uniform(-0.005, 0.005))
        hist_units = fund["units"] * random.uniform(0.98, 1.02)
        hist_nav_pu = hist_nav / hist_units
        change = (hist_nav / (base_nav * (1 + 0.01 * max(0, i - 1))) - 1) * 100 if i > 0 else 0
        wsh.cell(i + 2, 1, hist_date).number_format = 'DD/MM/YYYY'
        wsh.cell(i + 2, 2, round(hist_nav, 2)).number_format = '#,##0.00'
        wsh.cell(i + 2, 3, round(hist_nav_pu, 4)).number_format = '#,##0.0000'
        wsh.cell(i + 2, 4, round(hist_units, 0)).number_format = '#,##0'
        wsh.cell(i + 2, 5, round(change, 4)).number_format = '0.0000%'

    for col, w in zip("ABCDEFGHIJ", [14,12,32,10,18,18,14,18,16,20]):
        set_col_width(ws, col, w)
    wb.save(output_path)
    print(f"  ✓ Created {output_path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# FORMAT C — GENERIC / VARIED
# ══════════════════════════════════════════════════════════════════════════════
def create_generic_file(fund: dict, output_path: Path):
    wb = openpyxl.Workbook()
    sheet_name = fund.get("sheet_name", "NAV Summary")
    ws = wb.active
    ws.title = sheet_name

    # Intentionally varied layout per fund
    idx = [f["code"] for f in FUNDS].index(fund["code"])

    if fund["status"] == "missing_nav":
        # OMICRON015: NAV field is blank
        ws["A1"] = fund["name"]
        ws["B3"] = "Valuation Date"
        ws["C3"] = VALDATE
        ws["B5"] = "Net Asset Value (USD)"
        ws["C5"] = None  # ← BLANK — intentional error
        ws["B6"] = "NOTE: Preliminary report — NAV pending confirmation"
        ws["B6"].font = Font(color="C0392B", italic=True)

    elif fund["status"] == "broken_sheet":
        # NU013: sheet is "NAV_Summary" (underscore) not "NAV Summary"
        ws["A1"] = fund["name"]
        ws["A3"] = "Report Date"
        ws["B3"] = VALDATE
        ws["A5"] = "Total Net Asset Value"
        ws["B5"] = fund["nav"]
        ws["A6"] = "Currency"
        ws["B6"] = fund["currency"]

    else:
        # Normal generic funds — each has slightly different layout
        nav_date = VALDATE
        ws["A1"] = fund["name"]
        ws["A1"].font = Font(bold=True, size=13)

        # Some use "Report Date", some use "As Of", some use "Valuation Date"
        date_labels = ["Report Date", "As Of", "Statement Date", "Period End"]
        nav_labels  = ["Total NAV", "Fund Total", "Net Asset Value", "NAV Total"]
        ws["A3"] = date_labels[idx % 4]
        ws["B3"] = nav_date
        ws["B3"].number_format = 'DD-MMM-YYYY' if idx % 2 == 0 else 'MM/DD/YYYY'
        ws["A4"] = "Fund Code"
        ws["B4"] = fund["code"]
        ws["A5"] = nav_labels[idx % 4]
        ws["B5"] = fund["nav"]
        ws["B5"].number_format = '#,##0.00'
        ws["A6"] = "Currency"
        ws["B6"] = fund["currency"]
        ws["A7"] = "Units"
        ws["B7"] = fund["units"]
        ws["B7"].number_format = '#,##0'
        ws["A8"] = "NAV / Unit"
        ws["B8"] = round(fund["nav"] / fund["units"], 4) if fund["nav"] else None
        ws["A9"] = "Management Fee"
        ws["B9"] = fund["mgmt_fee_rate"]
        ws["B9"].number_format = '0.00%'
        ws["A10"] = "Administrator"
        ws["B10"] = fund["admin"]

    set_col_width(ws, "A", 26)
    set_col_width(ws, "B", 20)
    wb.save(output_path)
    print(f"  ✓ Created {output_path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# PDF REPORT SIMULATION
# (Plain text files that simulate what pdfplumber.extract_text() returns)
# ══════════════════════════════════════════════════════════════════════════════
PDF_TEMPLATE = """\
================================================================================
                      FUND ADMINISTRATOR SERVICES LTD
                    NET ASSET VALUE STATEMENT — OFFICIAL
================================================================================

Fund Name:          {fund_name}
Fund Code:          {fund_code}
Administrator:      {admin}
Valuation Date:     {nav_date}
Report Type:        Monthly NAV Statement
Currency:           {currency}

--------------------------------------------------------------------------------
SECTION 1: NET ASSET VALUE SUMMARY
--------------------------------------------------------------------------------

  Net Asset Value (Total):          {nav_formatted}
  Units Outstanding:                {units_formatted}
  Net Asset Value per Unit:         {nav_per_unit_formatted}
  Prior Period NAV:                 {prior_nav_formatted}
  Change in NAV:                    {change_formatted}
  Change % vs Prior Period:         {change_pct}

--------------------------------------------------------------------------------
SECTION 2: FEE SCHEDULE
--------------------------------------------------------------------------------

  Management Fee Rate:              {mgmt_fee_rate_pct}% per annum
  Management Fee (Monthly):         {mgmt_fee_monthly_formatted}
  Administration Fee Rate:          {admin_fee_rate_pct}% per annum
  Administration Fee (Monthly):     {admin_fee_monthly_formatted}
  Total Fees (Monthly):             {total_fees_formatted}

--------------------------------------------------------------------------------
SECTION 3: CERTIFICATION
--------------------------------------------------------------------------------

  This statement has been prepared by {admin} in accordance with the Fund's
  Offering Memorandum and constitutes the official NAV for the period ended
  {nav_date}.

  Authorised by: [SIGNED]
  Date Issued:   {issue_date}

================================================================================
  CONFIDENTIAL — FOR AUTHORISED RECIPIENTS ONLY
================================================================================
"""

def create_pdf_text_file(fund: dict, output_path: Path):
    """Create a .txt file that simulates pdfplumber.extract_text() output."""
    if fund["nav"] is None:
        nav = 0.0
    elif fund["status"] == "large_move":
        nav = fund["nav"] * 1.02  # ← slight difference vs Excel (creates a reconciliation exception)
    else:
        nav = fund["nav"]

    nav_per_unit = nav / fund["units"] if fund["units"] else 0
    prior = fund.get("nav_prior", nav * 0.99)
    change = nav - prior
    change_pct = (change / prior * 100) if prior else 0
    mgmt_monthly = nav * fund["mgmt_fee_rate"] / 12
    admin_monthly = nav * fund["admin_fee_rate"] / 12

    content = PDF_TEMPLATE.format(
        fund_name=fund["name"],
        fund_code=fund["code"],
        admin=fund["admin"],
        nav_date=VALDATE.strftime("%d %B %Y"),
        currency=fund["currency"],
        nav_formatted=f"{fund['currency']} {nav:>20,.2f}",
        units_formatted=f"{fund['units']:>24,.2f}",
        nav_per_unit_formatted=f"{fund['currency']} {nav_per_unit:>20,.4f}",
        prior_nav_formatted=f"{fund['currency']} {prior:>20,.2f}",
        change_formatted=f"{fund['currency']} {change:>+20,.2f}",
        change_pct=f"{change_pct:>+10.4f}%",
        mgmt_fee_rate_pct=f"{fund['mgmt_fee_rate']*100:.3f}",
        mgmt_fee_monthly_formatted=f"{fund['currency']} {mgmt_monthly:>20,.2f}",
        admin_fee_rate_pct=f"{fund['admin_fee_rate']*100:.3f}",
        admin_fee_monthly_formatted=f"{fund['currency']} {admin_monthly:>20,.2f}",
        total_fees_formatted=f"{fund['currency']} {mgmt_monthly+admin_monthly:>20,.2f}",
        issue_date=(VALDATE + timedelta(days=2)).strftime("%d %B %Y"),
    )
    output_path.write_text(content, encoding="utf-8")
    print(f"  ✓ Created {output_path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# SQLITE DATABASE (portable SQL Server replacement)
# ══════════════════════════════════════════════════════════════════════════════
def create_sqlite_database(db_path: Path):
    conn = sqlite3.connect(db_path)
    cur  = conn.cursor()

    # fund_master
    cur.execute("""
    CREATE TABLE IF NOT EXISTS fund_master (
        fund_code TEXT PRIMARY KEY,
        fund_name TEXT,
        currency TEXT,
        strategy TEXT,
        manager TEXT,
        administrator TEXT,
        inception_date TEXT,
        is_active INTEGER,
        nav_tolerance REAL,
        expected_nav_day TEXT
    )""")

    # daily_nav (matches what SQL Server view would return)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS daily_nav (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fund_code TEXT,
        nav_date TEXT,
        total_nav REAL,
        units_outstanding REAL,
        nav_per_unit REAL,
        currency TEXT,
        source TEXT,
        loaded_at TEXT
    )""")

    # nav_history
    cur.execute("""
    CREATE TABLE IF NOT EXISTS nav_history (
        fund_code TEXT,
        nav_date TEXT,
        total_nav REAL,
        nav_per_unit REAL,
        units_outstanding REAL,
        PRIMARY KEY (fund_code, nav_date)
    )""")

    # positions
    cur.execute("""
    CREATE TABLE IF NOT EXISTS positions (
        fund_code TEXT,
        trade_date TEXT,
        security_id TEXT,
        security_name TEXT,
        asset_class TEXT,
        quantity REAL,
        price REAL,
        market_value REAL,
        currency TEXT,
        pct_nav REAL
    )""")

    # Insert fund master
    for f in FUNDS:
        cur.execute("""
        INSERT OR REPLACE INTO fund_master VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            f["code"], f["name"], f["currency"], f["strategy"], f["manager"],
            f["admin"], "2015-01-01", 1, 1000.0, "T+1"
        ))

    # Insert today's NAV — SQL Server has SAME data as Excel (for reconciliation)
    # For DELTA004 (large_move), deliberately put a DIFFERENT value to create an exception
    for f in FUNDS:
        nav_in_sql = f["nav"]
        if f["status"] == "large_move":
            nav_in_sql = f["nav"] * 0.999  # SQL has slightly different value → reconciliation exception
        if nav_in_sql is None:
            continue
        nav_pu = nav_in_sql / f["units"] if f["units"] else None
        cur.execute("""
        INSERT INTO daily_nav (fund_code, nav_date, total_nav, units_outstanding,
                               nav_per_unit, currency, source, loaded_at)
        VALUES (?,?,?,?,?,?,?,?)
        """, (
            f["code"], VALDATE.isoformat(), nav_in_sql, f["units"],
            round(nav_pu, 4) if nav_pu else None,
            f["currency"], "SQL_WAREHOUSE", datetime.now().isoformat()
        ))

    # Insert 12 months of history
    for f in FUNDS:
        if f["nav"] is None:
            continue
        base = f["nav"] * 0.85
        for i in range(12):
            hist_date = VALDATE - timedelta(days=30 * (11 - i))
            hist_nav  = base * (1 + 0.01 * i + random.uniform(-0.003, 0.003))
            hist_pu   = hist_nav / f["units"]
            cur.execute("""
            INSERT OR IGNORE INTO nav_history VALUES (?,?,?,?,?)
            """, (f["code"], hist_date.isoformat(), round(hist_nav, 2),
                  round(hist_pu, 4), f["units"]))

    conn.commit()
    conn.close()
    print(f"  ✓ Created {db_path.name} with {len(FUNDS)} funds, 12 months history")


# ══════════════════════════════════════════════════════════════════════════════
# REFERENCE MASTER
# ══════════════════════════════════════════════════════════════════════════════
def create_fund_master(output_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fund Master"

    headers = ["Fund Code","Fund Name","Currency","Strategy","Manager",
               "Administrator","Format","Is Active","NAV Tolerance","Expected Day"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = header_font()
        c.fill = navy_fill()
        c.alignment = Alignment(horizontal="center")

    fmt_map = {"citco": "Format A — Citco", "ssc": "Format B — SS&C", "generic": "Format C — Generic"}
    for ri, f in enumerate(FUNDS, 2):
        row = [f["code"], f["name"], f["currency"], f["strategy"], f["manager"],
               f["admin"], fmt_map[f["admin_fmt"]], "Yes", 1000.0, "T+1"]
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)

    for col, w in zip("ABCDEFGHIJ", [12,35,10,18,28,22,22,10,14,10]):
        set_col_width(ws, col, w)
    wb.save(output_path)
    print(f"  ✓ Created {output_path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# CONFIG FILES
# ══════════════════════════════════════════════════════════════════════════════
def create_configs(config_dir: Path, data_dir: Path):
    settings = {
        "TENANT_ID":        "YOUR-TENANT-ID-HERE",
        "CLIENT_ID":        "YOUR-CLIENT-ID-HERE",
        "CLIENT_SECRET":    "YOUR-CLIENT-SECRET-HERE",
        "MAILBOX_EMAIL":    "fundops@yourcompany.com",
        "SHAREPOINT_SITE":  "https://yourcompany.sharepoint.com/sites/FundOps",
        "INPUT_FOLDER_CITCO":   str(data_dir / "input" / "format_a_citco"),
        "INPUT_FOLDER_SSC":     str(data_dir / "input" / "format_b_ssc"),
        "INPUT_FOLDER_GENERIC": str(data_dir / "input" / "format_c_generic"),
        "INPUT_FOLDER_PDF":     str(data_dir / "input" / "pdf_reports"),
        "BROKEN_FOLDER":        str(data_dir / "input" / "broken"),
        "REFERENCE_FOLDER":     str(data_dir / "reference"),
        "SQL_DB_PATH":          str(data_dir / "sql" / "funddb.sqlite"),
        "REPORTS_FOLDER":       str(data_dir.parent / "reports"),
        "LOGS_FOLDER":          str(data_dir.parent / "logs"),
        "NAV_TOLERANCE":        1000.0,
        "LARGE_MOVE_THRESHOLD": 0.05,
        "COLLECTION_SHEET_CITCO": "NAV Summary",
        "COLLECTION_SHEET_SSC":   "Fund_Performance",
        "NOTIFICATION_EMAIL":     "fundops-alerts@yourcompany.com",
        "VALUATION_DATE":         VALDATE.isoformat(),
        "_NOTE": "Replace TENANT_ID, CLIENT_ID, CLIENT_SECRET with real values from Azure Portal"
    }
    (config_dir / "settings.json").write_text(json.dumps(settings, indent=2), encoding="utf-8")

    exception_rules = {
        "_description": "Exception rules for the fund NAV validation engine. Add rules here — no code change needed.",
        "rules": [
            {"name": "NAV_MISSING",      "field": "NAVValue",    "condition": "is_null",   "severity": "CRITICAL", "action": "Contact fund administrator immediately"},
            {"name": "NAV_ZERO",         "field": "NAVValue",    "condition": "eq_zero",   "severity": "CRITICAL", "action": "Data error — verify with fund admin"},
            {"name": "NAV_NEGATIVE",     "field": "NAVValue",    "condition": "negative",  "severity": "CRITICAL", "action": "Impossible value — reject and investigate"},
            {"name": "DATE_FUTURE",      "field": "NAVDate",     "condition": "future_date","severity":"HIGH",     "action": "Verify valuation date with administrator"},
            {"name": "DATE_STALE",       "field": "NAVDate",     "condition": "stale_3d",  "severity": "MEDIUM",  "action": "Confirm fund is not on extended settlement"},
            {"name": "NAV_LARGE_MOVE",   "field": "NAVChange_Pct","condition":"gt_5_pct",  "severity": "HIGH",    "action": "Review pricing sources and market events"},
            {"name": "FEE_RATE_HIGH",    "field": "MgmtFeeRate", "condition": "gt_0025",   "severity": "MEDIUM",  "action": "Verify against IMA — may be performance fee period"},
            {"name": "UNITS_CHANGE",     "field": "UnitsChange_Pct","condition":"gt_10_pct","severity":"MEDIUM",  "action": "Confirm subscription/redemption activity"},
            {"name": "RECON_EXCEPTION",  "field": "Variance",    "condition": "gt_tolerance","severity":"HIGH",   "action": "Reconcile difference between PDF and Excel sources"}
        ]
    }
    (config_dir / "exception_rules.json").write_text(json.dumps(exception_rules, indent=2), encoding="utf-8")

    pipeline_config = {
        "pipeline_name": "FundOps Daily NAV Pipeline",
        "run_time": "08:00",
        "retry_attempts": 3,
        "retry_delay_seconds": 60,
        "email_on_success": True,
        "email_on_failure": True,
        "steps": [
            {"step": 1, "name": "outlook_agent",       "script": "scripts/day1/lab4_outlook_agent.py",     "enabled": True},
            {"step": 2, "name": "duplicate_check",     "script": "scripts/day1/lab7_duplicate_detector.py","enabled": True},
            {"step": 3, "name": "collection_citco",    "script": "scripts/day1/lab2_collection_agent.py",  "enabled": True, "format": "citco"},
            {"step": 4, "name": "collection_ssc",      "script": "scripts/day1/lab2_collection_agent.py",  "enabled": True, "format": "ssc"},
            {"step": 5, "name": "dynamic_discovery",   "script": "scripts/day1/lab3_discovery_agent.py",   "enabled": True},
            {"step": 6, "name": "pdf_extraction",      "script": "scripts/day2/lab8_pdf_extraction.py",    "enabled": True},
            {"step": 7, "name": "reconciliation",      "script": "scripts/day2/lab9_reconciliation.py",    "enabled": True},
            {"step": 8, "name": "sql_reconciliation",  "script": "scripts/day2/lab10_sql_integration.py",  "enabled": True},
            {"step": 9, "name": "exception_engine",    "script": "scripts/day2/lab12_exception_engine.py", "enabled": True},
            {"step":10, "name": "validation_engine",   "script": "scripts/day2/lab13_validation_engine.py","enabled": True},
            {"step":11, "name": "sharepoint_upload",   "script": "scripts/day2/lab14_sharepoint.py",       "enabled": False, "note": "Requires Graph API credentials"},
            {"step":12, "name": "audit_report",        "script": "scripts/day1/lab6_audit_framework.py",   "enabled": True},
        ]
    }
    (config_dir / "pipeline_config.json").write_text(json.dumps(pipeline_config, indent=2), encoding="utf-8")

    gitignore = "config/settings.json\nlogs/\n*.pyc\n__pycache__/\n*.egg-info/\n"
    (config_dir.parent / ".gitignore").write_text(gitignore, encoding="utf-8")

    print(f"  ✓ Created settings.json, exception_rules.json, pipeline_config.json, .gitignore")


# ══════════════════════════════════════════════════════════════════════════════
# HELPER: GENERATE SAMPLE POSITIONS
# ══════════════════════════════════════════════════════════════════════════════
SECURITIES = [
    ("AAPL US", "Apple Inc", "Equity"),
    ("MSFT US", "Microsoft Corp", "Equity"),
    ("AMZN US", "Amazon.com Inc", "Equity"),
    ("US912810TM06", "US Treasury 4.5% 2033", "Fixed Income"),
    ("US38141GXZ20", "Goldman Sachs 3.8% 2026", "Fixed Income"),
    ("GLD US", "SPDR Gold ETF", "Commodity"),
    ("EURUSD FWD", "EUR/USD Forward 3M", "FX Forward"),
    ("SPX INDEX", "S&P 500 Index Future", "Future"),
]

def _generate_positions(fund: dict, count: int):
    nav = fund["nav"] or 10_000_000
    rows = []
    remaining_pct = 100.0
    for i, (sec_id, sec_name, asset_class) in enumerate(SECURITIES[:count]):
        pct = round(random.uniform(8, 20) if i < count - 1 else remaining_pct, 2)
        remaining_pct -= pct
        mv = nav * (pct / 100)
        price = round(random.uniform(50, 500), 2)
        qty = round(mv / price, 0)
        rows.append([sec_id, sec_name, asset_class, qty, price, round(mv, 2), fund["currency"], pct / 100])
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# VERIFICATION REPORT — so participants can check their scripts
# ══════════════════════════════════════════════════════════════════════════════
def create_expected_output(output_path: Path):
    rows = []
    for f in FUNDS:
        rows.append({
            "FundCode": f["code"],
            "FundName": f["name"],
            "AdminFormat": f["admin_fmt"].upper(),
            "ExpectedNAV": f["nav"],
            "ExpectedDate": (TODAY + timedelta(days=1)).isoformat() if f["status"] == "future_date" else VALDATE.isoformat(),
            "Currency": f["currency"],
            "ExpectedStatus": f["status"].upper(),
            "ShouldExtractSuccessfully": "NO" if f["status"] in ("broken_sheet", "missing_nav") else "YES",
            "Note": {
                "ok": "Normal — should extract cleanly",
                "future_date": "NAVDate = tomorrow — should trigger DATE_FUTURE exception",
                "large_move": "NAV moved >8% — should trigger NAV_LARGE_MOVE exception. PDF has slightly different value → reconciliation exception",
                "broken_sheet": "Sheet name is NAV_Summary (underscore) not 'NAV Summary' — should log error",
                "missing_nav": "NAV cell is blank — should log CRITICAL exception",
            }[f["status"]]
        })
    df = pd.DataFrame(rows)
    df.to_excel(output_path, index=False)
    print(f"  ✓ Created {output_path.name} (verification reference for participants)")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN — CREATE EVERYTHING
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("\n" + "="*60)
    print("  FUND OPERATIONS WORKSHOP — DATA GENERATOR")
    print(f"  Root: {ROOT}")
    print(f"  Valuation Date: {VALDATE}")
    print("="*60 + "\n")

    # Create folder structure
    folders = [
        ROOT / "data" / "input" / "format_a_citco",
        ROOT / "data" / "input" / "format_b_ssc",
        ROOT / "data" / "input" / "format_c_generic",
        ROOT / "data" / "input" / "pdf_reports",
        ROOT / "data" / "input" / "broken",
        ROOT / "data" / "reference",
        ROOT / "data" / "sql",
        ROOT / "config",
        ROOT / "logs",
        ROOT / "reports" / "daily",
        ROOT / "reports" / "monthly",
        ROOT / "reports" / "exceptions",
        ROOT / "scripts" / "day1",
        ROOT / "scripts" / "day2",
    ]
    for f in folders:
        f.mkdir(parents=True, exist_ok=True)
    print("✅ Folder structure created\n")

    # Generate Excel files
    print("📊 Creating Fund Excel Files...")
    citco_dir   = ROOT / "data" / "input" / "format_a_citco"
    ssc_dir     = ROOT / "data" / "input" / "format_b_ssc"
    generic_dir = ROOT / "data" / "input" / "format_c_generic"

    for f in FUNDS:
        fname = f"{f['code']}_{f['name'].replace(' ', '_')[:20]}_NAV.xlsx"
        if f["admin_fmt"] == "citco":
            create_citco_file(f, citco_dir / fname)
        elif f["admin_fmt"] == "ssc":
            create_ssc_file(f, ssc_dir / fname)
        else:
            create_generic_file(f, generic_dir / fname)

    print("\n📄 Creating PDF Text Simulations...")
    pdf_dir = ROOT / "data" / "input" / "pdf_reports"
    # Create PDFs for a subset of funds (for reconciliation lab)
    pdf_funds = [f for f in FUNDS if f["admin_fmt"] in ("citco", "ssc")]
    for f in pdf_funds:
        fname = f"{f['code']}_Statement_{VALDATE.strftime('%Y%m%d')}.txt"
        create_pdf_text_file(f, pdf_dir / fname)

    print("\n🗄️  Creating SQLite Database (SQL Server simulation)...")
    create_sqlite_database(ROOT / "data" / "sql" / "funddb.sqlite")

    print("\n📋 Creating Reference Files...")
    create_fund_master(ROOT / "data" / "reference" / "fund_master.xlsx")

    # Expected output for verification
    create_expected_output(ROOT / "data" / "reference" / "EXPECTED_OUTPUT_verification.xlsx")

    print("\n⚙️  Creating Config Files...")
    create_configs(ROOT / "config", ROOT / "data")

    # Folder structure summary
    print("\n" + "="*60)
    print("  ✅ ALL DATA CREATED SUCCESSFULLY")
    print("="*60)
    print(f"""
FOLDER STRUCTURE:
{ROOT}/
├── data/
│   ├── input/
│   │   ├── format_a_citco/     ← 5 Citco-format Excel files (ALPHA–EPSILON)
│   │   ├── format_b_ssc/       ← 5 SS&C-format Excel files (ZETA–KAPPA)
│   │   ├── format_c_generic/   ← 5 Generic-format Excel files (LAMBDA–OMICRON)
│   │   │                          ↑ includes 2 broken files for error handling labs
│   │   └── pdf_reports/        ← 10 .txt files (simulate pdfplumber output)
│   ├── reference/
│   │   ├── fund_master.xlsx    ← all 15 funds with metadata
│   │   └── EXPECTED_OUTPUT_verification.xlsx ← what your script should produce
│   └── sql/
│       └── funddb.sqlite       ← SQLite DB (portable SQL Server replacement)
├── config/
│   ├── settings.json           ← all paths, tolerances, credentials
│   ├── exception_rules.json    ← configurable exception rules
│   ├── pipeline_config.json    ← full pipeline orchestration config
│   └── (settings.json is in .gitignore — never commit credentials!)
├── logs/                       ← empty, populated by your scripts
├── reports/
│   ├── daily/
│   ├── monthly/
│   └── exceptions/
└── scripts/
    ├── day1/                   ← Labs 1-7
    └── day2/                   ← Labs 8-15

DELIBERATE ERRORS (for teaching error handling):
  NU013    — sheet name is 'NAV_Summary' (underscore) not 'NAV Summary'
  OMICRON015 — NAV cell is blank (missing value)
  BETA002  — NAV date = tomorrow (future date — exception flag)
  DELTA004 — NAV moved >8% vs prior (large move — exception + recon difference)

NEXT STEP:
  Open VS Code → File → Open Folder → {ROOT}
  Then start Lab 1!
""")

if __name__ == "__main__":
    random.seed(42)
    main()
